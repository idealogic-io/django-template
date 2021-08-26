import datetime

from django.contrib.auth import authenticate
from django.core.mail import EmailMessage
from rest_framework import serializers, pagination
from rest_framework.exceptions import AuthenticationFailed
import openpyxl
from . import (
    models, utils, tasks
)
from django.conf import settings
from core.services import sms
from django.shortcuts import redirect


class SubscriptionSerializer(serializers.ModelSerializer):

    class Meta:
        model = models.Subscription
        fields = ('id', 'name', 'price')

class UserSubscriptionNotes(serializers.ModelSerializer):
    class Meta:
        model = models.UserSubscriptionNotes
        fields = ('id', 'text', 'created_at', 'created_by')
    def save(self, **kwargs):
        note = models.UserSubscriptionNotes.objects.create(
            plan_id=self.validated_data.get('plan_id', ''),
            text=self.validated_data.get('text', ''),
            created_by=self.context['request'].user.name,
        )

class UserSubscriptionSerializer(serializers.ModelSerializer):
    plan = SubscriptionSerializer()
    notes = UserSubscriptionNotes(many=True)
    class Meta:
        model = models.UserSubscription
        fields = ('id', 'plan', 'expiry', 'notes')

class UserCreateSubscriptionSerializer(serializers.Serializer):
    plan_id = serializers.IntegerField()
    expiry = serializers.DateTimeField()

class RegistrationModelSerializer(serializers.ModelSerializer):
    class Meta:
        model = models.Registration
        fields = ('code', 'key', 'phone')
        extra_kwargs = {
            'code': {
                   'write_only': True
                  }
        }

class PasswordRestoreModelSerializer(serializers.ModelSerializer):
    class Meta:
        model = models.Registration
        fields = ('code', 'key', 'phone')
        extra_kwargs = {
            'code': {
                'write_only': True
            }
        }

class RestoreValidator(serializers.Serializer):
    email = serializers.EmailField()

    def validate(self, data):
        if not models.User.objects.filter(email=data['email']).exists():
            raise serializers.ValidationError(
                'incorrect email'
            )

        return data
        # lookup = sms.lookup(data['phone'])

        # print(lookup)
        # if lookup:
        #     return data
        # raise serializers.ValidationError(
        #     'Incorrect phone number or email'
        # )

    def proceed(self, validated_data):
        code = utils.generate_phone_secret()
        user = models.User.objects.get(email=self.validated_data['email'])
        user.pass_reset_code = code
        now = datetime.datetime.now()
        now_plus_10 = now + datetime.timedelta(minutes = 5)
        user.pass_reset_date = now_plus_10
        user.save()
        # key = utils.generate_registration_token()
        # phone = self.validated_data.get('phone','')

        # reg_key = models.PasswordRestore.objects.create(
        #     code=code,
        #     key=key,
        #     phone=phone.replace(" ", ""),
        #     email=self.validated_data.get('email','')
        # )

        # tasks.send_restore_password_code.delay(reg_key.phone, reg_key.code)
        email = self.validated_data.get('email', '')

        # tasks.send_restore_email.delay(email, code)
        text = "Hello! \n" \
               "Received password restore request\n" \
               "Use code below to proceed:\n" + \
                code

        email = EmailMessage(
            "Password restore",
            text,
            'info@djangoboilerplate.com',
            [self.validated_data.get('email')]
        )
        email.send(fail_silently=False)

        return "email"
        # return PasswordRestoreModelSerializer(reg_key).data

class ApproveRestoreValidator(serializers.Serializer):
    code = serializers.CharField(max_length=255)
    # key = serializers.CharField(max_length=255)
    password = serializers.CharField(max_length=255)

    def validate(self, data):
        user = models.User.objects.get(pass_reset_code=self.initial_data.get('code', ''))
        if user.pass_reset_date <= datetime.datetime.now():
            raise serializers.ValidationError(
                'Your code is outdated'
            )
        return data

    def validate_password(self, password):
        if not utils.is_password_secure(password):
            raise serializers.ValidationError(
                'Password is not secure'
            )
        return password

    def proceed(self, validated_data):
        # restore = models.PasswordRestore.objects.get(code=self.validated_data['code'])
        user = models.User.objects.get(
            pass_reset_code=self.validated_data.get('code')
        )
        user.set_password(self.validated_data.get('password', ''))
        user.pass_reset_code = None
        user.pass_reset_date = None
        user.save()
        print(user.email)
        return validated_data

class UserValidator(serializers.Serializer):
    name = serializers.CharField(max_length=255)
    email = serializers.EmailField(max_length=255)
    password = serializers.CharField(max_length=255)
    phone_number = serializers.CharField(max_length=20)

    def validate_password(self, password):
        if not utils.is_password_secure(password):
            raise serializers.ValidationError(
                'Password is not secure'
            )
        return password

    def validate_email(self, email):
        if models.User.objects.filter(email=email.lower()).exists():
            raise serializers.ValidationError(
                'Email already exists'
            )

    def validate_phone_number(self, number):
        lookup = sms.lookup(number)
        if lookup:
            return number
        raise serializers.ValidationError(
            'Incorrect phone number'
        )

    def proceed(self, validated_data):
        code = utils.generate_phone_secret()
        key = utils.generate_registration_token()
        phone = validated_data.get('phone_number')
        reg_key = models.Registration.objects.create(
            code=code,
            key=key,
            phone=phone.replace(" ", "")
        )
        tasks.send_very_sms.delay(reg_key.phone, reg_key.code)
        return reg_key


class UserDeviceSerializer(serializers.ModelSerializer):
    class Meta:
        model = models.UserDevice
        fields = ('id', 'user', 'one_signal_id', 'device_id')
        extra_kwargs = {
            'user': {
                'read_only': True
            },
            'one_signal_id': {
                'read_only': True
            }
        }

    def save(self, **kwargs):
        device = models.UserDevice.objects.create(
            user_id=self.context['request'].user.id,
            device_id=self.validated_data.get('device_id', '')
        )
        return device


class FakeNotes(serializers.Serializer):
    text = serializers.CharField()


class CreateUserManuallySerializer(serializers.ModelSerializer):
    email = serializers.CharField(max_length=50)
    password = serializers.CharField(max_length=16)
    name = serializers.CharField(max_length=50, required=False)
    phone_number = serializers.CharField(max_length=12, required=False)

    class Meta:
        model = models.User
        fields = (
             'email', 'password', 'name', 'phone_number'
        )

    def create(self):
        user = models.User.objects.create(
            email=self.validated_data['email'].lower()
        )
        user.set_password(self.validated_data.get('password', 'Qwer1234'))
        user.phone_number = self.validated_data.get('phone_number', '').replace(" ", "")
        user.is_phone_verified = True
        user.name = self.validated_data.get('name', '')
        user.save()
        text = "Hello! \n" \
               "Welcome to Django Boilerplate!\n" \
               "Now you can login with Your Credentials:\n" \
               "Email: " + self.validated_data.get('email') + \
               "Password: " + self.validated_data.get('password', 'Qwer1234')

        email = EmailMessage(
            "Your created Account on Django Boilerplate",
            text,
            'info@djangoboilerplate.com',
            [self.validated_data.get('email')]
        )
        email.send(fail_silently=False)
        #tasks.send_welcome_email.delay(user.id)
        return user


class UserSerializer(serializers.ModelSerializer):
    devices = UserDeviceSerializer(many=True)
    user_plan = UserSubscriptionSerializer()

    class Meta:
        model = models.User
        fields = (
            'id', 'email', 'role', 'is_phone_verified', 'phone_number',
            'created_at', 'last_login', 'name', 'devices', 'auth_token', 'is_superuser', 'callout_fee', 'user_plan'
        )
        extra_kwargs = {
            'callout_fee': {
                'read_only': True
            }
        }


class SafeUserSerializer(serializers.ModelSerializer):
    devices = UserDeviceSerializer(many=True)
    user_plan = UserSubscriptionSerializer()

    class Meta:
        model = models.User
        fields = (
            'id', 'email', 'role', 'is_phone_verified', 'phone_number',
            'created_at', 'last_login', 'name', 'devices','user_plan', 'is_superuser', 'callout_fee'
        )
        extra_kwargs = {
            'callout_fee': {
                'read_only': True
            }
        }


class UpdateUser(serializers.Serializer):
    email = serializers.EmailField(required=False)
    phone_number = serializers.CharField(max_length=255, required=False)
    name = serializers.CharField(max_length=255, required=False)
    password = serializers.CharField(max_length=255, required=False, allow_blank=True)

    def validate_phone_number(self, number):
        lookup = sms.lookup(number)
        if lookup:
            return number
        raise serializers.ValidationError(
            'Incorrect phone number, please use international phone format'
        )

    def validate_password(self, password):
        if self.initial_data.get('password', ''):
            if not utils.is_password_secure(password):
                raise serializers.ValidationError(
                    'Password is not secure'
                )
        return password

    def validate_email(self, email):

        email_user = models.User.objects.filter(email=email.lower())
        if email_user.exists():
            if email_user.get().id == self.context.get("request").user.id:
                pass
            if email_user.get().id != self.context.get("request").user.id:
                raise serializers.ValidationError(
                    {"password": "User with this email already registered"}
                )

        return email

    def save_update(self, user):

        if self.validated_data.get('password', ''):
            user.set_password(self.validated_data.get('password', ''))
        user.phone_number = self.validated_data.get('phone_number', user.phone_number).replace(" ", "")
        user.email = self.validated_data.get('email', user.email).lower()
        user.name = self.validated_data.get('name', user.name)

        user.save()
        return user


class LoginSerializer(serializers.Serializer):
    email = serializers.EmailField()
    password = serializers.CharField(max_length=255)

    def validate(self, attrs):
        self.user = authenticate(**{
            'email': attrs['email'],
            'password': attrs['password'],
        })

        if self.user is None \
                or not self.user.is_active \
                or not self.user.is_email_verified:
            raise AuthenticationFailed(
                'No active account found with the given credentials'
            )
        data = UserSerializer(self.user).data
        return data

class UpdateUserSerializer(serializers.Serializer):
    def write(self):
        model = models.User
        queryset = model.objects.all()

        for user in queryset:
            if user.user_plan is None:
                plan = UserSubscriptionSerializer().create({'plan_id': 1, 'expiry': None })
                user.user_plan = plan
                user.save()


class AdminLoginSerializer(serializers.Serializer):
    email = serializers.EmailField()
    password = serializers.CharField(max_length=255)

    def validate(self, attrs):
        self.user = authenticate(**{
            'email': attrs['email'],
            'password': attrs['password'],
        })

        if self.user is None or self.user.role is not 'operator' and not self.user.is_staff:
            raise AuthenticationFailed(
                'No active account found with the given credentials'
            )
        data = UserSerializer(self.user).data
        return data


class RegisterSerializer(serializers.ModelSerializer):
    key = serializers.CharField(max_length=255)

    class Meta:
        model = models.User
        fields = ('email', 'password', 'name', 'phone_number', 'key')
        extra_kwargs = {
            'password': {
                'write_only': True
            }
        }

    def validate_password(self, password):
        if not utils.is_password_secure(password):
            raise serializers.ValidationError(
                'Password is not secure'
            )
        return password

    # def validate_key(self, token):
    #     try:
    #         obj = models.Registration.objects.get(key=token)
    #         if obj.phone == self.initial_data['phone_number'].replace(" ", "") and obj.code_approved:
    #             return token
    #         raise serializers.ValidationError(
    #             "Wrong phone number"
    #         )
    #     except Exception:
    #         raise serializers.ValidationError(
    #             "Wrong key"
    #         )

    def save(self, **kwargs):
        user = models.User.objects.create(
            email=self.validated_data['email'].lower()
        )
        user.set_password(self.validated_data.get('password', ''))
        user.phone_number = self.validated_data.get('phone_number', '').replace(" ", "")
        user.is_phone_verified = False
        user.name = self.validated_data.get('name', '')
        user.role = 'customer'
        user.save()
        text = "Hello! \n" \
               "Welcome to Django Boilerplate!\n" \
               "Please, confirm your account creation by clicking a link below:\n" \
               + settings.BASE_URL + 'api/v1/auth/users/verify_email/?secret=' + str(user.uuid) + ''

        email = EmailMessage(
            "Welcome in the Django Boilerplate",
            text,
            'info@djangoboilerplate.com',
            [self.validated_data['email'].lower()]
        )
        email.send(fail_silently=False)
        # plan = UserSubscriptionSerializer().create({'plan_id': 1, 'expiry': None })
        # user.user_plan = plan
        # user.welcome_email = True
        # user.save()
        # tasks.send_welcome_email.delay(user.id)
        return user


class OperatorRegisterSerializer(serializers.ModelSerializer):

    class Meta:
        model = models.User
        fields = ('name', 'email', 'phone_number', 'password')
        extra_kwargs = {
            'password': {
                'write_only': True
            }
        }

    def validate_password(self, password):
        if not utils.is_password_secure(password):
            raise serializers.ValidationError(
                'Password is not secure'
            )
        return password

    def save(self, **kwargs):

        user = models.User.objects.create(
            email=self.validated_data['email'].lower()
        )
        user.set_password(self.validated_data.get('password', ''))
        user.phone_number = self.validated_data.get('phone_number', '').replace(" ", "")
        user.is_active = True
        user.is_staff = True
        user.role = 'operator'
        user.is_phone_verified = True
        user.name = self.validated_data.get('name', '')
        user.save()
        return user

    def set_new(self, instance):
        user = instance
        user.set_password(self.validated_data.get('password', user.password))
        user.save()
        return user

class UserUpdateSerializer(serializers.ModelSerializer):
    plan = UserCreateSubscriptionSerializer()
    notes = FakeNotes(many=True)
    class Meta:
        model = models.User
        fields = ('name', 'email', 'plan', 'notes', 'phone_number', 'created_at')

    def set_new(self, instance):
        user = instance
        notes = self.validated_data.pop('notes', [])
        user.name = self.validated_data.get('name', user.name)
        user.email = self.validated_data.get('email', user.email).lower()
        user.phone_number = self.validated_data.get('phone_number', user.phone_number).replace(" ", "")
        user.created_at = self.validated_data.get('created_at', user.created_at)

        if self.validated_data.get('plan', user.user_plan) is not user.user_plan:
            plan = UserSubscriptionSerializer().create(self.validated_data.get('plan', user.user_plan))
            if plan.plan.id == 1:
                plan.expiry = None
                plan.save()
            user.user_plan = plan
        [p.setdefault('plan_id', user.user_plan.id) for p in notes]
        UserSubscriptionNotes(many=True).create(notes)
        user.save()
        return user

class OperatorUpdateSerializer(serializers.ModelSerializer):

    class Meta:
        model = models.User
        fields = ('name', 'email')

    def set_new(self, instance):
        user = instance
        user.name = self.validated_data.get('name', user.name)
        user.email = self.validated_data.get('email', user.email).lower()
        user.phone_number = self.validated_data.get('phone_number', user.phone_number).replace(" ", "")
        user.save()
        return user


class SetOperatorPasswordSerializer(serializers.Serializer):

    password = serializers.CharField(
        max_length=255
    )
    confirm_password = serializers.CharField(
        max_length=255
    )

    @staticmethod
    def validate_password(password):
        if not utils.is_password_secure(password):
            raise serializers.ValidationError(
                'Password is not secure'
            )
        return password

    @staticmethod
    def validate_confirm_password(password):
        if not utils.is_password_secure(password):
            raise serializers.ValidationError(
                'Password is not secure'
            )
        return password

    def validate(self, attrs):
        if attrs.get('password') != \
                attrs.get('confirm_password'):
            raise serializers.ValidationError('Password are not same')
        return attrs


class RestorePasswordEmail(serializers.Serializer):
    email = serializers.EmailField()


class SetPasswordSerializer(serializers.Serializer):

    old_password = serializers.CharField(
        max_length=255
    )
    password = serializers.CharField(
        max_length=255
    )
    confirm_password = serializers.CharField(
        max_length=255
    )

    def validate_old_password(self, old_password):
        request = self.context.get("request")
        if request and hasattr(request, "user"):
            user = request.user
            if not user.check_password(old_password):
                raise serializers.ValidationError(
                    'Password is incorrect'
                )
            return user

    @staticmethod
    def validate_password(password):
        if not utils.is_password_secure(password):
            raise serializers.ValidationError(
                'Password is not secure'
            )
        return password

    @staticmethod
    def validate_confirm_password(password):
        if not utils.is_password_secure(password):
            raise serializers.ValidationError(
                'Password is not secure'
            )
        return password

    def validate(self, attrs):
        if attrs.get('password') != \
                attrs.get('confirm_password'):
            raise serializers.ValidationError('Password are not same')
        return attrs


class RestorePassword(serializers.Serializer):
    sign = serializers.CharField(
        max_length=255
    )
    password = serializers.CharField(
        max_length=255
    )
    confirm_password = serializers.CharField(
        max_length=255
    )

    @staticmethod
    def validate_sign(sign):
        user_id = utils.parse_secret(sign)
        user = utils.get_or_none(models.User, pk=user_id)
        if not user:
            raise serializers.ValidationError(
                'Url is wrong'
            )
        return user

    @staticmethod
    def validate_password(password):
        if not utils.is_password_secure(password):
            raise serializers.ValidationError(
                'Password is not secure'
            )
        return password

    @staticmethod
    def validate_confirm_password(password):
        if not utils.is_password_secure(password):
            raise serializers.ValidationError(
                'Password is not secure'
            )
        return password

    def validate(self, attrs):
        if attrs.get('password') != \
                attrs.get('confirm_password'):
            raise serializers.ValidationError('Password are not same')
        return attrs


class StandardResultsSetPagination(pagination.PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100

class FileSerializer(serializers.Serializer):
    type = serializers.ChoiceField(choices=(('users', 'users'), ('callouts', 'callouts'), ('suppliers', 'suppliers')))
    all_bool = serializers.BooleanField()
    date_from = serializers.DateField(required=False)
    date_to = serializers.DateField(required=False)

class BulkUsersCreateSerializer(serializers.Serializer):
    csv = serializers.FileField()

    def write(self, request):
        csv_file = request.FILES['csv']
        wb_obj = openpyxl.load_workbook(csv_file)

        sheet_obj = wb_obj.active
        m_row = sheet_obj.max_row
        max_col = sheet_obj.max_column

        for iq in range(2, m_row + 1):
            user_data = dict()
            for i in range(1, max_col + 1):
                cell_obj = sheet_obj.cell(row = iq, column = i)
                cell_name = sheet_obj.cell(row = 1, column = i)
                user_data[cell_name.value] = cell_obj.value
            print(user_data)
            already_exists = models.User.objects.filter(
                email=user_data['email'].lower()
            ).exists()
            if not already_exists and user_data['email'] is not None:
                user = models.User.objects.create(
                    email=user_data['email'].lower()
                )
                user.set_password('offline!user')
                user.phone_number = user_data['phone_number'].replace(" ", "")
                user.is_phone_verified = True
                user.name = user_data['name']
                user.created_at = user_data['created_at']
                user.save()
                general_plan = models.Subscription.objects.get(id=user_data['plan'])
                expiry = user_data['expiry']
                if general_plan == 1:
                    expiry = None
                plan = models.UserSubscription.objects.create(
                    expiry= expiry,
                    plan=general_plan
                )
                user.user_plan = plan
                user.save()


class UpdateUserManuallySerializer(serializers.ModelSerializer):
    plan = UserCreateSubscriptionSerializer()
    notes = FakeNotes(many=True)

    class Meta:
        model = models.User
        fields = (
            'email', 'role', 'phone_number',
            'name', 'plan', 'notes'
        )
        extra_kwargs = {
            'user_plan': {
                'read_only': True
            },
            'email': {
                'read_only': True
            }
        }

    def validate_phone_number(self, phone):
        lookup = sms.lookup(phone)
        if lookup:
            return phone
        raise serializers.ValidationError(
            'Incorrect phone number'
        )

    def update(self, instance, validated_data):
        user = instance
        plan = self.initial_data.pop('plan', '')
        if models.User.objects.filter(email=self.initial_data.get('email', user.email)).exclude(id=user.id).exists():
            raise serializers.ValidationError(
                'user with this email already exists'
            )
        user.email = self.initial_data.get('email', user.email)
        notes = self.initial_data.pop('notes', [])

        user.phone_number = self.validated_data.get('phone_number', user.phone_number).replace(" ", "")
        user.name = self.validated_data.get('name', user.name)

        user.created_at = self.initial_data.get('created', user.created_at)

        if user.user_plan is None or plan['plan_id'] != user.user_plan.plan_id:
            if user.user_plan is not None:
                user.user_plan.delete()
            user_plan = UserSubscriptionSerializer().create(plan)
            if user_plan.plan.id == 1:
                user_plan.expiry = None
                user_plan.save()
            user.user_plan = user_plan
        [p.setdefault('plan_id', user.user_plan.id) for p in notes]
        if len(notes) > 0:
            UserSubscriptionNotes(many=True).create(notes)
        user.save()
        user.plan = user.user_plan
        user.notes = notes
        return user

